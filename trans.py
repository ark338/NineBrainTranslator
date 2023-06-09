import os
import openpyxl
import json
import time
import re
from easy_gpt_utils import gpt, embedding, vector_database
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging
import sys

class PrintHandler(logging.StreamHandler):
    def emit(self, record):
        print(self.format(record))

formatter = logging.Formatter('%(asctime)s [%(levelname)s] [%(name)s] [%(funcName)s:%(lineno)d] - %(message)s')
console_handler = PrintHandler(sys.stdout)
console_handler.setFormatter(formatter)
logger = logging.getLogger()
logger.setLevel(logging.INFO)
logger.setLevel(logging.DEBUG)
logger.addHandler(console_handler)

# True to use openai False to use Azure
use_openai = False
glossary_token_limit = 300

def make_query(title, label, text):
    #return f"The texts to be translated are under the category of [{title}] and [{label}].\n The texts to be translated are:[{text}]"
    return f"The texts to be translated are:['''{text}''']"

def get_context(gpt_instance, query_text, max_retries = 3):
    # query glossaies from vector database and set it as context
    if use_openai:
        # use openai
        embe_instance = embedding.Embedding()
    else :
        # use azure
        embe_instance = embedding.Embedding(
            model="model-text-embedding-ada-002", 
            api_type="azure", 
            api_base = "https://ninebot-rd-openai-1.openai.azure.com/",
            api_version = "2022-12-01")
    
    pinecorn_instance = vector_database.Pinecone(index = 'segway-knowledge-base', environment='asia-southeast1-gcp')

    ret = None
    for i in range(max_retries):
        try:
            logger.debug("start embe")
            embe = embe_instance.get_raw_embedding(query_text)
            logger.debug("embe finished")
            # TODO: buggy here, token usage did not calculated correctly
            ret = pinecorn_instance.query_meta(namespace = vector_database.NamesSpaces.Glossary.value, threshold = 0.85, vector = embe, top_k = 5)
            logger.debug(f"pinecorn finished, ret: {ret}")
            break
        except Exception as e:
            logger.error(f"get_context error: {e}")
            # to workaround pinecone bug, sleep and retry
            time.sleep(1)
            continue

    context = ""       
    if ret is not None and len(ret) > 0:
        token_count = 0
        context = "Here is the glossary, please refer to the glossary for translation. If the meaning is the same, please use the translation in the glossary directly:" + "\n".join(item['metadata']['content'] for item in ret if (token_count := token_count + gpt_instance.num_tokens_from_string(item['metadata']['content'])) < glossary_token_limit)

    return context

def trans(gpt_instance, query_text, history=None):
    # Implement your translation API logic here.
    # For example, using Google Translate API:
    if (history):
        gpt_instance.set_history(history)

    context = get_context(gpt_instance, query_text)
    translations = gpt_instance.query([context], query_text)

    logger.debug(f"query finished query_text: {query_text}, translations = {translations}")
    return translations

def log_error(row_number, row_data, error, translations_str=None):
    error_str = f"Error at row {row_number}: {error}\nInput data: {row_data}\nReturned data: {translations_str}\n"
    logger.error(error_str)
    with open("log.txt", "a") as log_file:
        log_file.write(error_str)

# 这个函数用于将一个列表中的内容进行翻译，返回一个翻译后的列表
# 其中，列表的第一个元素是分类，第二个元素是标题，第三个元素是url(没有用到)，第四个元素是正文
# 返回的列表中，前四个元素是原来的内容，后面的元素是翻译后的内容，分别是波兰语(pl)，德语(de)，俄语(ru)，法语(fr)，韩语(ko)，荷兰语(nl)，日语(ja)，土耳其语(tr)，西班牙语(es)，意大利语(it)，越南语(vi)
# 如果翻译失败，那么就会返回None
# 由于GPT-3的限制，每次翻译的内容不能超过300个token，会调用长句翻译，将正文分成多个部分，然后分别翻译，最后再将翻译后的内容合并起来，因此可以处理大文本
# 但是GPT-3在进行大文本翻译的时候非常非常慢
# 但是由于gpt类中split_text函数的限制，如果一个单句的token数量超过1800（预设值，可修改），则会出现错误返回None
# input args:
#  row_number: 行号,用来表示在input excel中的行位置，这里仅用户打印
#  row_data: 一行输入数据，是一个列表，包含了分类，标题，url，正文
#  min_interval: 最小间隔，如果翻译时间小于min_interval，那么会sleep，这是由于GPT调用时往往会在一定时间内有次数限制，如果短时间内调用次数过多，会返回错误
#  retries: 重试次数，如果翻译失败，会重试retries次，如果仍然失败，那么就会返回None
def process_row(row_number, row_data, target_languages, progress_callback=None, min_interval=1, retries=3, enable_gpt4 = False):
    start_time = time.time()

    # 168是一个经验值，如果token数量小于168，那么就调用短句翻译，否则调用长句翻译
    # 猜测大于170的时候，GPT-3会出现错误，实际中遇到过token达到170时不出错，到171的时候就会出现错误
    # 猜测应该是要求单句token数量 * 11国语言 + prompt的token数量 < 2048
    # 当语言数量不确定的时候，采用1850/语言数量 作为阈值(gpt4使用3700)
    # TODO: 2023/05/11 buggy here, 在增加了glossary之后，token的使用量计算不正确，需要重新计算
    # 2021/05/11 临时方案，再减去一个glossary_token_limit / 2

    if (enable_gpt4 is None or enable_gpt4 == False or enable_gpt4 == "false"):
        if (use_openai):
            # use openai
            gpt_instance = gpt.GPT(model = "gpt-3.5-turbo")
        else:
            # use azure
            gpt_instance = gpt.GPT(
                model="model-gpt-35-turbo-0301", 
                api_type="azure", 
                api_base = "https://ninebot-rd-openai-1.openai.azure.com/",
                api_version = "2023-03-15-preview",
                api_key = os.getenv("AZURE_API_KEY"))
        token_limit = 1850
    else:
        gpt_instance = gpt.GPT(model = "gpt-4")
        token_limit = 3700

    tokens = gpt_instance.num_tokens_from_string(row_data[3])
    fast_token_limit = int((token_limit - glossary_token_limit / 2) / len(target_languages))
    logger.info(f"processing row: {row_number} tokens: {tokens} fast_token_limit: {fast_token_limit}")
    if progress_callback:
        progress_callback(f"processing string: {row_number} tokens: {tokens}")

    try:
        if (tokens <= fast_token_limit):
            return process_row_shot(gpt_instance, row_number, row_data, target_languages, progress_callback, retries)     
        else:
            return process_row_long(gpt_instance, row_number, row_data, target_languages, progress_callback, retries)
    finally:
        end_time = time.time()
        elapsed_time = end_time - start_time

        if elapsed_time < min_interval:
            time.sleep(min_interval - elapsed_time)

        logger.info(f"row {row_number} processed in {elapsed_time} seconds")
        if progress_callback:
            progress_callback(f"string {row_number} processed in {elapsed_time} seconds")

def process_row_long(gpt_instance, row_number, row_data, target_languages, progress_callback=None, retries=3):
    title, label, _, paragraph = row_data
    if not paragraph:
        return row_data + (None,) * len(target_languages)
    translations = []
    
    translated_texts = [None] * len(target_languages)

    for lan_index, language in enumerate(target_languages):
        gpt_instance.set_system_prompt(
        f'''You are a senior translator and work for Segway-Ninebot which has products such as emoped, e-bikes, e-scooters, go-karts, segways, off-road vehicles, and unicycles. They also use an App to operate the vehicles, get assistance, and communicate with users. I will provide you with the text to be translated, and you need to translate the text into [{language}]. The format is as follows:
        Translated text
        ''')
        gpt_instance.set_post_prompt("")
        gpt_instance.set_use_history(False)

        texts = gpt_instance.split_text(paragraph)
        logger.debug(f"Row {row_number} split: {len(texts)} language: {language} texts: {texts}")
        translated_text = ""
        for index, text in enumerate(texts):
            query_text = make_query(title, label, text)
            logger.info(f"processing row {row_number} language: {language}({lan_index + 1}/{len(target_languages)}) split: {index + 1}/{len(texts)}")
            if (progress_callback):
                progress_callback(f"processing string {row_number} language: {language}({lan_index + 1}/{len(target_languages)}) split: {index + 1}/{len(texts)}")
            break_label = False
            for attempt in range(retries):
                try:
                    translations_text = trans(gpt_instance, query_text)
                    logger.debug(f"Row {row_number} translations_str: {translations_text}")
                    translated_text = translated_text + translations_text
                    break
                except Exception as e:
                    if attempt < retries - 1:
                        logger.warning(f"Row {row_number} Exception, attempt{attempt}, error:{e}")
                        # sleep 1 seconds to avoid rate limit bug like pinecorn had
                        time.sleep(1)
                        continue
                    log_error(row_number, row_data, e)
                    break_label = True
                    break

            if break_label:
                translated_text = ""
                logger.error(f"Row {row_number} language:{language}({lan_index + 1}/{len(target_languages)} failed to translate, break")
                break

        translations.append(translated_text)
    translated_texts = translations
    return row_data + tuple(translated_texts)

def process_row_shot(gpt_instance, row_number, row_data, target_languages, progress_callback=None, retries=3):
    title, label, _, text = row_data
    if not text:
        return row_data + (None,) * len(target_languages)

    translated_texts = [None] * len(target_languages)

    gpt_instance.set_system_prompt(
    #'''你是Segway Ninebot的高级翻译人员，Segway Ninebot生产两轮电动车、滑板车、卡丁车、平衡车、机甲战车、独轮车等硬件产品，并使用App来操作车辆，获取帮助，和用户交流等。
    #我将给你要翻译的文本，你需要将文本翻译为波兰语(pl)，德语(de)，俄语(ru)，法语(fr)，韩语(ko)，荷兰语(nl)，日语(ja)，土耳其语(tr)，西班牙语(es)，意大利语(it)，越南语(vi)。
    #请输出11种语言的翻译，以json数组的格式输出，例子如下：
    #[
    #{"lan":翻译语言1的简称，例如ko,
    #"txt":翻译后的文本},
    #{"lan":翻译语言2的简称，例如ja,
    #"txt":翻译后的文本}
    #]"
    #注意，当翻译后的文本中有双引号时会影响json解析，请确保在引号前添加json的转义字符。其他影响json解析的符号也需要添加转义字符。在翻译后的文本中不要出现换行符，否则会影响json解析。
    #'''
    '''You are a senior translator and work for Segway-Ninebot which has products such as emoped, e-bikes, e-scooters, go-karts, segways, off-road vehicles, and unicycles. They also use an App to operate the vehicles, get assistance, and communicate with users. I will provide you with the text to be translated, and you need to translate the text into ''' + str(target_languages) + '''. Use json array format to output the translation. For example:
    [
    {"lan":abbreviation of the first language, for instance pl,
    "txt":translated text},
    {"lan":abbreviation of the second language, for instance de,
    "txt":translated text}
    ]
    Instrction: When there are double quotes in the translated text, it will affect the json parsing. Please make sure to add the escape character of json before the quotes. Other symbols that affect json parsing also need to add escape characters. Do not use line breaks in the translated text, otherwise it will affect json parsing.
    '''
    )
    gpt_instance.set_post_prompt("")
    gpt_instance.set_use_history(False)
    history = None

    query_text = make_query(title, label, text)

    translations_str = ""
    for attempt in range(retries):
        try:
            translations_str = trans(gpt_instance, query_text, history)
            logger.debug(f"Row {row_number} translations_str: {translations_str}")
            json_pattern = re.compile(r'\[\s*\{[\s\S]*\}\s*\]')
            json_string = json_pattern.search(translations_str).group()
            translations = json.loads(json_string)
            translated_texts = [t["txt"] for t in translations]
            break
        except Exception as e:
            if attempt < retries - 1:
                logger.warning(f"Row {row_number} Exception, attempt {attempt}, error:{e}, raw text: {translations_str}")
                history = [
                    {'role':'user', 'content':query_text},
                    {'role':'assistant', 'content':translations_str}
                    ]
                query_text = '''
                    When I try to parse the returned JSON, I encountered an error:
                    {e}
                    Please make corrections.
                    '''
                # The json format is wrong, it may be:\n1. special characters, such as no escape character before the quotation mark, or there may be extra characters. In this case you should fix it.\n2. There are more than one json array, In this case you should no split the text to be translated.\n please check and re-output.
                gpt_instance.set_use_history(True)
                # sleep 1 seconds to avoid rate limit bug like pinecorn had
                time.sleep(1)
                continue
            log_error(row_number, row_data, e, translations_str)
            break

    return row_data + tuple(translated_texts)

def process_excel(input_file, output_file, min_row = 2, max_row = None, target_languages = None, progress_callback=None, enable_gpt4 = False):    
    input_wb = openpyxl.load_workbook(input_file)
    input_sheet = input_wb.active

    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active

    if max_row == None:
        max_row = input_sheet.max_row

    # Copy headers from input sheet to output sheet
    headers = [cell for cell in next(input_sheet.iter_rows(min_row=1, max_row=1, max_col=4, values_only=True))] + target_languages
    output_sheet.append(headers)

    # Process each row in the input sheet using ThreadPoolExecutor
    input_data = list(input_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=4, values_only=True))

    row_count = 0
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(process_row, row_number, row_data, target_languages, enable_gpt4=enable_gpt4): row_data for row_number, row_data in enumerate(input_data, start=min_row)}

        for future in futures:
            row_count = row_count + 1
            progress_callback(f"{row_count}/{max_row} rows processed")
            output_data = future.result()
            output_sheet.append(output_data)
            output_wb.save(output_file)

    if progress_callback:
        progress_callback('All translation finished.')

    input_wb.close()
    output_wb.close()

def main(input_file, output_file, min_row, max_row):
    target_languages = ["Polish (pl)", "German (de)", "Russian (ru)", "French (fr)", "Korean (ko)", "Dutch (nl)", "Japanese (ja)", "Turkish (tr)", "Spanish (es)", "Italian (it)", "Vietnamese (vi)"]
    process_excel(input_file, output_file, min_row, max_row, target_languages=target_languages)

if __name__ == "__main__":
    input_file = "longinput.xlsx"
    output_file = "longout.xlsx"
    min_row = 2
    max_row = 55
    main(input_file, output_file, min_row, max_row)